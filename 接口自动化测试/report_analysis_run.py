# Author:Feiling Peng
from function.add_lib import addLibs
addLibs()
import os
import sys
import requests
import re
from openpyxl import Workbook, load_workbook
import json



if __name__ == '__main__':
    # *生成的excel在report目录下
    # *测试报告(html_files_path)：必须为pytest - html生成的文档, 可设置多个（直接文件名，无需压缩）
    # *原测试记录excel文档(excel_file)：用于更新用，没有可不上传，会创建新的文档
    # *测试结果excel名称关键字(file_name_key)：可自定义新生成测试记录excel名称，默认值为“自动化测试”
    # *html报告是否为同一个环境(is_one)：True--同一个环境,False--多个环境
    #
    # *excel问题记录更新规则：获取html中失败、错误用例的接口异常信息，如果excel中不存在该接口问题记录，则新增；如果存在则更新数据（不更新日期）
    # *excel已有问题记录“验证结果”更新规则：获取html中失败、错误用例的接口异常信息，如果excel中不存在该接口问题记录，则更新为“是”。

    # *所以会存在这种情况：
    # 原excel中有接口B的记录，但本次因为前一步报错导致没有跑到接口B,那么report中不会有接口B的错误，excel中接口B的“验证结果”字段会被更新为是。
    # 考虑的处理方法是：用例运行时，将运行过的接口记录在log中（代码已经写过并提交了）。本报告分析中再读取下log，如果原excel中的接口即不在report异常接口又不在log中，就不更新“验证结果”字段（还未做）。
    args = sys.argv
    test_user = 'baidi'  # 测试人姓名
    html_num = 11  # 要选择的html报告份数
    file_name_key = ''  # 生成的文件名
    excel_file = 'old'  # 要写入的excel路径,当值为old时，代表从report目录取最近的一次名称包含'自动化测试.xlsx'的excel
    save_num = 2 # 保留报告的份数
    file_name = '' # 报告文件重命名 mysql_job.xlsx
    fail_counts = {}  #汇总失败接口影响的用例范围个数

    if len(args) > 1:
        html_num = int(args[1])
        test_user = args[2]
        if len(args) > 3:
            excel_file = args[3]
        if len(args) > 4:
            file_name_key = args[4]
        if len(args) > 5:
            save_num = args[5]
        if len(args) > 6:
            file_name = args[6]

    html_dir_path = test_report_dir  # html报告所在文件目录
    # 当excel_file值为old时，代表从report目录取最近的一次名称包含'自动化测试.xlsx'的excel
    if excel_file == 'old':
        old_excels = choose_latest_file(html_dir_path, 1, '自动化测试.xlsx')
        excel_file = old_excels[0] if old_excels else ''
    # 当excel_file有值，但系统中不存在时，设置为空值
    if not os.path.exists(excel_file):
        excel_file=''
    html_files_path = choose_latest_file(html_dir_path, html_num)  # html报告路径列表
    is_one = True  # 多份html报告是否是同一个测试环境
    if not html_files_path or '' in html_files_path:
        raise ValueError('请检查报告文件路径！')
    if not excel_file:
        new_file_path = os.path.join(html_dir_path, file_name_key + '自动化测试.xlsx')
        wb = Workbook()
        wb.save(new_file_path)
    else:
        excel_file_path = excel_file
        new_file_path = copy_file(excel_file_path, file_name_key + '自动化测试', test_report_dir)
    # 报告分析
    wb = load_workbook(new_file_path)
    sheetsnames = wb.sheetnames
    for sheet in sheetsnames:
        if '当日问题记录' in sheet:
            ws = wb[sheet]
            wb.remove(ws)
            wb.save(new_file_path)
        if '未运行用例模块' in sheet:
            ws = wb[sheet]
            wb.remove(ws)
            wb.save(new_file_path)

    pages = {}  #{环境:[报告对象]}
    for path in html_files_path:
        page = ReportAnalysis(path, test_user, new_file_path, is_one)
        page.analysis()
        if page.result_dic['url'] not in pages:
            pages[page.result_dic['url']] = []
        pages[page.result_dic['url']].append(page)
    # if len(pages) < 2 and not is_one:
    #     raise ValueError('上传的html文件属于同一个环境,请重新设置is_one!')
    # 报告写入
    url_result = {} #{环境:{'passed':12312,'skipped':123123,'failed':0.0,'error':0.0,'run_test_num':7589,} 得到各个结果的统计总和
    url_info = {}  # url_exit_info
    url_exception = {}  # exception_url
    url_time = {}
    modle_data = {}
    for url, value in pages.items():
        modle_data[url] = {}
        if url not in url_result:
            url_result[url] = {k: 0.00 for k, v in value[0].result_dic.items() if k != 'url'}
            url_result[url]['run_test_num'] = 0
        for page in value:
            url_result[url].update(
                {k: v + page.result_dic[k] for k, v in url_result[url].items() if k in page.result_dic})
            url_result[url]['run_test_num'] += page.run_test_num
            page.write_info_to_excel()
            modle_data[url].update(page.result_dic_module)
            # 本次所有抛出异常的接口信息处理
            if url not in url_info:
                url_info[url] = page.url_exit_info
                url_exception[url] = page.exception_url
            else:
                url_info[url].update(page.url_exit_info)
                url_exception[url].update(page.exception_url)
            if url not in url_time:
                url_time[url] = {'start': set(), 'end': set()}
                url_time[url]['start'].add(page.start_time)
                url_time[url]['end'].add(page.report_time)
            else:
                url_time[url]['start'].add(page.start_time)
                url_time[url]['end'].add(page.report_time)
            fail_count_case = page.count
            fail_counts.update(fail_count_case)
            # print(fail_counts)
    for k, v in url_time.items():
        v['start'] = sorted(v['start'])[0]
        v['end'] = sorted(v['end'], reverse=True)[0]
    # 报告运行结果sheet页写入、旧问题记录更新
    for url, res in url_result.items():
        page = ReportAnalysis(html_files_path[0], test_user, new_file_path, is_one)

        res['ratio_run'] = round(res['passed'] / res['run_test_num'],
                                 4) if res['run_test_num'] else 0
        all = res['passed'] + res['failed'] + res['error']
        res['ratio_run_and_error'] = round(res['passed'] / all, 4) if all else 0
        page.run_test_num = res['run_test_num']
        page.result_dic = res
        page.result_dic['url'] = url
        page.url_exit_info = url_info[url]
        page.exception_url = url_exception[url]
        page.start_time = url_time[url]['start']
        page.report_time = url_time[url]['end']
        page.write_run_result()
        page.update_old_row()
        page.print_info()


    last_page = list(pages.values())[-1][-1]
    last_page.write_exception_sheet(file_path=new_file_path)  # 服务不可用
    last_page.write_exception_sheet('SQL异常模块', 'SQLServerException', file_path=new_file_path)
    last_page.write_exception_sheet('待解决问题模块', '', True, file_path=new_file_path)

    #处理报告产生的历史数据
    all_excels = choose_files(html_dir_path, save_num, '自动化测试.xlsx')
    all_html_files = choose_files(html_dir_path, save_num*html_num)  # html报告路径列表
    if all_excels:
        for file_excel in all_excels:
            os.remove(file_excel)
    if all_html_files:
        for file_html in all_html_files:
            os.remove(file_html)

    wb = load_workbook(new_file_path)
    ReportAnalysis.get_sheets(wb, new_file_path)
    #html_files_path = html_files_path[len(html_files_path)-1]
    for html_file in html_files_path:
        print(f'当前解析的报告: {html_file}')
        wb = load_workbook(new_file_path)
        #ReportAnalysis.get_sheets(wb, new_file_path)
        ReportAnalysis.add_sheet_not_api_error(wb, new_file_path, html_file, test_user)
    '''
    # 统计当日问题记录接口影响的用例个数
    df = pd.read_excel(new_file_path, sheet_name='当日问题记录', header=None, engine="openpyxl")
    data = numpy.array(df)
    data = data.tolist()
    header = data[0]
    header.append('当前接口影响用例数')
    content = data[1:]
    new_content = []
    for i in content:
        for j in fail_counts:
            if ';' in j:
                api = j.split(';')[0]
            else:
                api = j
            if i[1] == api:
                i.append(len(fail_counts[j]))
                new_content.append(i)
    for item in new_content:
        if len(item) > 21:
            del item[21:]
    book = load_workbook(new_file_path)
    s1 = pd.DataFrame(new_content, columns=header)
    with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
        writer.book = book
        writer.sheets = {i.title: i for i in book.worksheets}
        s1.to_excel(writer, sheet_name="当日问题记录", index=False)
    '''

    # 发送给平台数据库
    try:
        import socket
        ip = socket.gethostbyname(socket.getfqdn(socket.gethostname()))
        url = 'http://192.168.60.66:8888/auth/report/analysis/import/db'
        # url = 'http://127.0.0.1:8888/auth/report/analysis/import/db'
        data = {"url_result": url_result, "url_time": url_time, "modle_data": modle_data, "report_dir": test_report_dir,
                "ip": ip}
        analysis_res = requests.post(url=url, data=json.dumps(data)).json()
        print("发送平台数据库，返回：",analysis_res)
        if analysis_res["code"] != 200:
            print("数据传给服务器数据库失败：", analysis_res["msg"])
    except Exception as e:
        print("数据传给服务器数据库失败：", e)

    '''
    # 将分析的后excel上传到django服务中并写入到数据库
    if file_name != '':
        files = [('file', (file_name, open(new_file_path, 'rb'), "text/txt"))]
        url = 'http://192.168.3.181:8000/uploadfile/'
        res = requests.post(url=url, files=files).json()
        if res['msg'] == '文件上传成功':
            url = 'http://192.168.3.181:8000/analysisexcel/'
            data = {'file_name': file_name}
            headers = {'Content-Type': 'application/json'}
            requests.post(url=url, json=data, headers=headers)
        else:
            raise RuntimeError('文件未上传到服务器')
    '''

    '''
    # 处理执行失败的模块
    utils_path = os.path.join(os.path.abspath(__file__))
    utils_path = utils_path.split('report_analysis_run')[0]
    final_moudle_result = []
    package_module_path = f'{utils_path}utils\package_module.yaml'
    with open(package_module_path, encoding='utf-8') as file:
        package_module = yaml.load(file.read(), Loader=yaml.FullLoader)
        df = pd.read_excel(new_file_path, sheet_name='未运行用例模块',
                           header=None, engine="openpyxl")
        data = numpy.array(df)
        data = data.tolist()
        header = data[0]
        content = data[1:]

        case_list = []
        for case in content:
            case_list.append(case[1])
        moudle_result = [content[0][0], '', '', '', '', '', '']
        for case_name in package_module:
            if case_name != 'a_initUser':
                if case_name not in case_list:
                    moudle_result.insert(1, case_name)
                    moudle_result.insert(2, package_module[case_name]['ch_name'])
                    moudle_result.insert(3, package_module[case_name]['user'])
                    moudle_result.append('当前模块下的用例未执行或跳过')
                    moudle_result.append(package_module[case_name]['auto_user'])
                    final_moudle_result.append(moudle_result)
                    moudle_result = [content[0][0], '', '', '', '', '', '']
        book = load_workbook(new_file_path)
        s1 = pd.DataFrame(final_moudle_result, columns=header)
        with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
            writer.book = book
            writer.sheets = {i.title: i for i in book.worksheets}
            s1.to_excel(writer, sheet_name="未运行用例模块", index=False)
    '''
    # 修改html报告属性，使其打开的时候所有结果都是关闭不撑开内容
    htmls = choose_latest_file(html_dir_path, 6)

    for i in htmls:
        os.chdir(html_dir_path)
        str1 = "else {extras.classList.add('collapsed');expandcollapse.classList.add('expander');}" \
               "elem.appendChild"
        with open(i, 'r', encoding='utf-8') as f:
            content = f.read()
            result = re.sub(r'else {\s+(.*?);\s+}\s+elem.appendChild', str1, content)
            new_name = i.split('.')[0]
            new_name = f'{new_name}备份.html'
            with open(new_name, 'w', encoding='utf-8') as f1:
                f1.write(result)
        os.remove(i)
        os.rename(new_name, i)
